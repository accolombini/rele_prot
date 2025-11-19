"""
Normalized Excel Exporter
Exporta dados normalizados para Excel individual (multi-sheet 3FN)
"""

from pathlib import Path
from typing import Dict, Any
from datetime import datetime

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill
    from openpyxl.utils import get_column_letter
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False


class NormalizedExcelExporter:
    """Exporta dados normalizados para Excel individual"""
    
    def __init__(self, output_dir: str = "outputs/norm_excel", logger=None):
        self.output_dir = Path(output_dir)
        self.output_dir.mkdir(parents=True, exist_ok=True)
        self.logger = logger
        
        if not OPENPYXL_AVAILABLE:
            if self.logger:
                self.logger.warning("openpyxl not available - Excel export disabled")
    
    def export_normalized(self, normalized_data: Dict[str, Any], base_filename: str) -> str:
        """
        Export normalized data to Excel workbook
        
        Args:
            normalized_data: Dictionary with relay_info, cts, vts, protections, parameters
            base_filename: Base name for output file
        
        Returns:
            Path to created Excel file
        """
        if not OPENPYXL_AVAILABLE:
            if self.logger:
                self.logger.warning("Skipping Excel export (openpyxl not available)")
            return None
        
        filename = f"{base_filename}_NORMALIZED.xlsx"
        filepath = self.output_dir / filename
        
        try:
            wb = Workbook()
            
            # Remove default sheet
            if 'Sheet' in wb.sheetnames:
                wb.remove(wb['Sheet'])
            
            # Create sheets
            self._create_summary_sheet(wb, normalized_data)
            self._create_cts_sheet(wb, normalized_data)
            self._create_vts_sheet(wb, normalized_data)
            self._create_protections_sheet(wb, normalized_data)
            self._create_parameters_sheet(wb, normalized_data)
            self._create_metadata_sheet(wb, normalized_data)
            
            # Save
            wb.save(filepath)
            
            if self.logger:
                relay_id = normalized_data['relay_info']['relay_id']
                self.logger.info(f"    ✓ Excel normalized: {filename} ({relay_id})")
            
            return str(filepath)
            
        except Exception as e:
            if self.logger:
                self.logger.error(f"Failed to export normalized Excel: {str(e)}")
            raise
    
    def _create_summary_sheet(self, wb: Workbook, data: Dict):
        """Create Summary sheet with relay info"""
        ws = wb.create_sheet("Summary", 0)
        relay_info = data['relay_info']
        
        # Title
        ws['A1'] = "RELAY SUMMARY - NORMALIZED DATA"
        ws['A1'].font = Font(bold=True, size=14)
        
        # Relay information
        row = 3
        info_fields = [
            ('Relay ID', 'relay_id'),
            ('Source File', 'source_file'),
            ('Manufacturer', 'manufacturer'),
            ('Model', 'model'),
            ('Barras Identifier', 'barras_identificador'),
            ('Config Date', 'config_date'),
            ('Frequency (Hz)', 'frequency_hz'),
            ('Software Version', 'software_version'),
            ('Processed At', 'processed_at')
        ]
        
        for label, key in info_fields:
            ws[f'A{row}'] = label
            ws[f'A{row}'].font = Font(bold=True)
            ws[f'B{row}'] = relay_info.get(key, '')
            row += 1
        
        # Statistics
        row += 2
        ws[f'A{row}'] = "STATISTICS"
        ws[f'A{row}'].font = Font(bold=True, size=12)
        row += 1
        
        stats = [
            ('CTs', len(data.get('cts', []))),
            ('VTs', len(data.get('vts', []))),
            ('Protection Functions', len(data.get('protections', []))),
            ('Total Parameters', len(data.get('parameters', [])))
        ]
        
        for label, value in stats:
            ws[f'A{row}'] = label
            ws[f'A{row}'].font = Font(bold=True)
            ws[f'B{row}'] = value
            row += 1
        
        # Auto-adjust column widths
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 40
    
    def _create_cts_sheet(self, wb: Workbook, data: Dict):
        """Create CTs sheet"""
        ws = wb.create_sheet("CTs")
        cts = data.get('cts', [])
        
        if not cts:
            ws['A1'] = "No CT data available"
            return
        
        # Headers
        headers = ['CT ID', 'Relay ID', 'Type', 'Primary (A)', 'Secondary (A)', 'Ratio', 'Usage']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(1, col, header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
            cell.font = Font(bold=True, color='FFFFFF')
        
        # Data
        for row_idx, ct in enumerate(cts, 2):
            ws.cell(row_idx, 1, ct.get('ct_id', ''))
            ws.cell(row_idx, 2, ct.get('relay_id', ''))
            ws.cell(row_idx, 3, ct.get('ct_type', ''))
            ws.cell(row_idx, 4, ct.get('primary_a', ''))
            ws.cell(row_idx, 5, ct.get('secondary_a', ''))
            ws.cell(row_idx, 6, ct.get('ratio', ''))
            ws.cell(row_idx, 7, ct.get('usage', ''))
        
        # Auto-adjust
        for col in range(1, 8):
            ws.column_dimensions[get_column_letter(col)].width = 15
    
    def _create_vts_sheet(self, wb: Workbook, data: Dict):
        """Create VTs sheet"""
        ws = wb.create_sheet("VTs")
        vts = data.get('vts', [])
        
        if not vts:
            ws['A1'] = "No VT data available"
            return
        
        # Headers
        headers = ['VT ID', 'Relay ID', 'Type', 'Primary (V)', 'Secondary (V)', 'Ratio']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(1, col, header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
            cell.font = Font(bold=True, color='FFFFFF')
        
        # Data
        for row_idx, vt in enumerate(vts, 2):
            ws.cell(row_idx, 1, vt.get('vt_id', ''))
            ws.cell(row_idx, 2, vt.get('relay_id', ''))
            ws.cell(row_idx, 3, vt.get('vt_type', ''))
            ws.cell(row_idx, 4, vt.get('primary_v', ''))
            ws.cell(row_idx, 5, vt.get('secondary_v', ''))
            ws.cell(row_idx, 6, vt.get('ratio', ''))
        
        # Auto-adjust
        for col in range(1, 7):
            ws.column_dimensions[get_column_letter(col)].width = 15
    
    def _create_protections_sheet(self, wb: Workbook, data: Dict):
        """Create Protections sheet"""
        ws = wb.create_sheet("Protections")
        prots = data.get('protections', [])
        
        if not prots:
            ws['A1'] = "No protection data available"
            return
        
        # Headers
        headers = ['Prot ID', 'Relay ID', 'ANSI Code', 'Function Name', 'Enabled', 
                  'Setpoint 1', 'Unit 1', 'Time Dial', 'Curve Type']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(1, col, header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
            cell.font = Font(bold=True, color='FFFFFF')
        
        # Data
        for row_idx, prot in enumerate(prots, 2):
            ws.cell(row_idx, 1, prot.get('prot_id', ''))
            ws.cell(row_idx, 2, prot.get('relay_id', ''))
            ws.cell(row_idx, 3, prot.get('ansi_code', ''))
            ws.cell(row_idx, 4, prot.get('function_name', ''))
            ws.cell(row_idx, 5, 'Yes' if prot.get('is_enabled') else 'No')
            ws.cell(row_idx, 6, prot.get('setpoint_1', ''))
            ws.cell(row_idx, 7, prot.get('unit_1', ''))
            ws.cell(row_idx, 8, prot.get('time_dial', ''))
            ws.cell(row_idx, 9, prot.get('curve_type', ''))
        
        # Auto-adjust
        for col in range(1, 10):
            width = 15 if col != 4 else 30
            ws.column_dimensions[get_column_letter(col)].width = width
    
    def _create_parameters_sheet(self, wb: Workbook, data: Dict):
        """Create Parameters sheet (all parameters)"""
        ws = wb.create_sheet("Parameters")
        params = data.get('parameters', [])
        
        if not params:
            ws['A1'] = "No parameter data available"
            return
        
        # Headers
        headers = ['Param ID', 'Relay ID', 'Section/Code', 'Parameter Name', 
                  'Value', 'Continuation Lines', 'Timestamp']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(1, col, header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
            cell.font = Font(bold=True, color='FFFFFF')
        
        # Data (limit to avoid Excel overload)
        max_rows = min(len(params), 10000)  # Safety limit
        for row_idx, param in enumerate(params[:max_rows], 2):
            ws.cell(row_idx, 1, param.get('param_id', ''))
            ws.cell(row_idx, 2, param.get('relay_id', ''))
            ws.cell(row_idx, 3, param.get('section_or_code', ''))
            ws.cell(row_idx, 4, param.get('parameter_name', ''))
            ws.cell(row_idx, 5, param.get('value', ''))
            ws.cell(row_idx, 6, param.get('continuation_lines', ''))
            ws.cell(row_idx, 7, param.get('timestamp', ''))
        
        if len(params) > max_rows:
            ws.cell(max_rows + 2, 1, f"Note: Showing {max_rows} of {len(params)} parameters")
        
        # Auto-adjust
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 20
        ws.column_dimensions['D'].width = 40
        ws.column_dimensions['E'].width = 30
        ws.column_dimensions['F'].width = 50
        ws.column_dimensions['G'].width = 20
    
    def _create_metadata_sheet(self, wb: Workbook, data: Dict):
        """Create Metadata sheet"""
        ws = wb.create_sheet("Metadata")
        
        ws['A1'] = "EXPORT METADATA"
        ws['A1'].font = Font(bold=True, size=14)
        
        row = 3
        ws[f'A{row}'] = "Export Date"
        ws[f'B{row}'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        row += 1
        
        ws[f'A{row}'] = "Format"
        ws[f'B{row}'] = "3NF (Third Normal Form)"
        row += 1
        
        ws[f'A{row}'] = "Pipeline Phase"
        ws[f'B{row}'] = "FASE 2 - Normalization"
        row += 1
        
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 40
