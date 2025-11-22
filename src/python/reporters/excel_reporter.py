"""
Exportador de relatórios em formato Excel com formatação
"""
import pandas as pd
from pathlib import Path
from typing import Optional, Dict
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from .base_reporter import BaseReporter


class ExcelReporter(BaseReporter):
    """Gera relatórios em formato Excel com formatação Petrobras"""
    
    def __init__(self, output_base_path: Optional[Path] = None):
        super().__init__(output_base_path)
        
        # Fontes
        self.header_font = Font(name='Arial', size=14, bold=True, color='FFFFFF')
        self.title_font = Font(name='Arial', size=11, bold=True, color='002366')
        self.data_font = Font(name='Arial', size=10)
        self.footer_font = Font(name='Arial', size=9, italic=True, color='666666')
        
        # Preenchimentos
        self.header_fill = PatternFill(start_color='002366', end_color='002366', fill_type='solid')
        self.title_fill = PatternFill(start_color='FFB81C', end_color='FFB81C', fill_type='solid')
        self.alt_row_fill = PatternFill(start_color='F0F0F0', end_color='F0F0F0', fill_type='solid')
        
        # Bordas
        thin_border = Side(style='thin', color='CCCCCC')
        self.border = Border(left=thin_border, right=thin_border, top=thin_border, bottom=thin_border)
        
        # Alinhamentos
        self.center_align = Alignment(horizontal='center', vertical='center')
        self.left_align = Alignment(horizontal='left', vertical='center')
    
    def export(
        self,
        df: pd.DataFrame,
        report_code: str,
        report_name: str,
        report_title: str,
        sheet_name: str = 'Relatório'
    ) -> Path:
        """
        Exporta DataFrame para Excel com formatação
        
        Args:
            df: DataFrame com os dados
            report_code: Código do relatório (ex: REL01)
            report_name: Nome descritivo (ex: fabricantes_reles)
            report_title: Título completo do relatório
            sheet_name: Nome da planilha
        
        Returns:
            Path do arquivo gerado
        """
        self.validate_dataframe(df)
        
        output_path = self.get_output_path(report_code, report_name, 'xlsx')
        
        wb = Workbook()
        ws = wb.active
        ws.title = sheet_name
        
        # Cabeçalho (linha 1)
        ws.merge_cells('A1:' + self._get_column_letter(len(df.columns)) + '1')
        header_cell = ws['A1']
        header_cell.value = f"{self.HEADER_SYMBOL}  {self.HEADER_TITLE}"
        header_cell.font = self.header_font
        header_cell.fill = self.header_fill
        header_cell.alignment = self.center_align
        ws.row_dimensions[1].height = 25
        
        # Título do relatório (linha 2)
        ws.merge_cells('A2:' + self._get_column_letter(len(df.columns)) + '2')
        title_cell = ws['A2']
        title_cell.value = report_title
        title_cell.font = self.title_font
        title_cell.alignment = self.center_align
        ws.row_dimensions[2].height = 20
        
        # Espaço (linha 3)
        ws.row_dimensions[3].height = 10
        
        # Dados (a partir da linha 4)
        start_row = 4
        
        # Determinar tamanhos de fonte baseado no número de colunas
        num_cols = len(df.columns)
        header_font_size = 9 if num_cols > 10 else 10
        data_font_size = 9 if num_cols > 10 else 10
        header_height = 60 if num_cols > 10 else 50
        
        # Cabeçalho dos dados
        for col_idx, column_name in enumerate(df.columns, 1):
            cell = ws.cell(row=start_row, column=col_idx)
            cell.value = column_name
            cell.font = Font(name='Arial', size=header_font_size, bold=True)
            cell.fill = self.title_fill
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = self.border
        
        # CORREÇÃO: Aumentar altura da linha de cabeçalho para acomodar quebras
        ws.row_dimensions[start_row].height = header_height
        
        # Linhas de dados
        for row_idx, row_data in enumerate(dataframe_to_rows(df, index=False, header=False), start_row + 1):
            max_line_count = 1  # Rastrear maior número de linhas na célula
            for col_idx, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.value = value
                cell.font = Font(name='Arial', size=data_font_size)
                # CORREÇÃO: Adicionar wrap_text=True para quebra automática
                cell.alignment = Alignment(
                    horizontal='left',
                    vertical='top',
                    wrap_text=True  # Habilita quebra de linha
                )
                cell.border = self.border
                
                # Contar linhas do texto para ajustar altura
                if value:
                    line_count = str(value).count('\n') + 1
                    max_line_count = max(max_line_count, line_count)
                
                # Zebrar linhas
                if (row_idx - start_row) % 2 == 0:
                    cell.fill = self.alt_row_fill
            
            # CORREÇÃO: Ajustar altura da linha baseada no conteúdo
            ws.row_dimensions[row_idx].height = max(18, max_line_count * 15)
        
        # CORREÇÃO: Ajustar largura das colunas - FORÇAR MÍNIMOS GENEROSOS
        for col_idx, column_name in enumerate(df.columns, 1):
            # Calcular comprimento máximo com multiplicador para fonte proporcional
            header_length = len(str(column_name).replace('\n', ''))
            data_max = df[column_name].astype(str).str.len().max() if not df.empty else 10
            calculated_width = max(header_length * 2.0, data_max * 1.8)  # Multiplicadores DOBRADOS
            
            # FORÇAR larguras mínimas generosas por tipo de coluna
            if 'Par' in str(column_name) and 'Crit' in str(column_name):
                col_width = max(calculated_width, 100)  # Mínimo 100
            elif 'Modelo' in str(column_name):
                col_width = max(calculated_width, 18)  # Mínimo 18 para "SEPAM S40"
            elif 'Tipo' in str(column_name) and 'Relé' in str(column_name):
                col_width = max(calculated_width, 16)  # Mínimo 16 para "PROT_TRAFO"
            elif 'Data' in str(column_name) and 'Config' in str(column_name):
                col_width = max(calculated_width, 12)  # Mínimo 12 para datas de 6 dígitos
            elif 'V_kV' in str(column_name):
                col_width = max(calculated_width, 10)  # Mínimo 10 para tensão
            elif 'SE' == str(column_name):
                col_width = max(calculated_width, 8)  # Mínimo 8 para SE
            elif 'Barra' in str(column_name):
                col_width = max(calculated_width, 10)  # Mínimo 10
            elif 'Fab' in str(column_name):
                col_width = max(calculated_width, 8)  # Mínimo 8 para GE, SNE
            elif 'Ver.' in str(column_name) and 'SW' in str(column_name):
                col_width = max(calculated_width, 20)  # AUMENTADO para 20 (Ver. SW com valores longos)
            elif 'Ver.' in str(column_name) and 'FW' in str(column_name):
                col_width = max(calculated_width, 8)  # REDUZIDO para 8 (Ver. FW só tem "None")
            elif 'Ver.' in str(column_name):
                col_width = max(calculated_width, 12)  # Outras versões: 12
            elif 'Prot' in str(column_name) or 'Tot' in str(column_name):
                col_width = max(calculated_width, 10)  # Mínimo 10 para contadores
            elif 'TP' in str(column_name) or 'Fonte' in str(column_name) or 'Conf' in str(column_name):
                col_width = max(calculated_width, 10)  # Mínimo 10
            else:
                col_width = max(calculated_width, 12)  # Padrão: mínimo 12
            
            # Aplicar limite máximo para não estourar página
            col_width = min(col_width, 100)
            
            ws.column_dimensions[self._get_column_letter(col_idx)].width = col_width
        
        # Rodapé (última linha + 2)
        footer_row = ws.max_row + 2
        footer = self.format_footer_text(report_title)
        
        ws.merge_cells(f'A{footer_row}:' + self._get_column_letter(len(df.columns)) + f'{footer_row}')
        footer_cell = ws[f'A{footer_row}']
        footer_cell.value = f"{footer['left']} | {footer['center']} | {footer['right']}"
        footer_cell.font = self.footer_font
        footer_cell.alignment = self.center_align
        
        wb.save(output_path)
        return output_path
    
    def export_multiple_sheets(
        self,
        sheets: Dict[str, pd.DataFrame],
        report_code: str,
        report_name: str,
        report_title: str
    ) -> Path:
        """
        Exporta múltiplas planilhas em um único arquivo Excel
        
        Args:
            sheets: Dict onde chave=nome_sheet, valor=DataFrame
            report_code: Código do relatório
            report_name: Nome descritivo
            report_title: Título completo
        
        Returns:
            Path do arquivo gerado
        """
        output_path = self.get_output_path(report_code, report_name, 'xlsx')
        
        wb = Workbook()
        wb.remove(wb.active)  # Remove planilha padrão
        
        for sheet_name, df in sheets.items():
            ws = wb.create_sheet(title=sheet_name)
            
            # Aplicar mesma formatação
            self._format_sheet(ws, df, report_title)
        
        wb.save(output_path)
        return output_path
    
    def _format_sheet(self, ws, df: pd.DataFrame, report_title: str):
        """Aplica formatação a uma planilha"""
        # Determinar tamanhos de fonte baseado no número de colunas
        num_cols = len(df.columns)
        header_font_size = 9 if num_cols > 10 else 10
        data_font_size = 9 if num_cols > 10 else 10
        header_height = 60 if num_cols > 10 else 50
        
        # Cabeçalho
        ws.merge_cells('A1:' + self._get_column_letter(len(df.columns)) + '1')
        header_cell = ws['A1']
        header_cell.value = f"{self.HEADER_SYMBOL}  {self.HEADER_TITLE}"
        header_cell.font = self.header_font
        header_cell.fill = self.header_fill
        header_cell.alignment = self.center_align
        ws.row_dimensions[1].height = 25
        
        # Título
        ws.merge_cells('A2:' + self._get_column_letter(len(df.columns)) + '2')
        title_cell = ws['A2']
        title_cell.value = report_title
        title_cell.font = self.title_font
        title_cell.alignment = self.center_align
        ws.row_dimensions[2].height = 25
        
        ws.row_dimensions[3].height = 10
        
        # Dados
        start_row = 4
        for col_idx, column_name in enumerate(df.columns, 1):
            cell = ws.cell(row=start_row, column=col_idx)
            cell.value = column_name
            cell.font = Font(name='Arial', size=header_font_size, bold=True)
            cell.fill = self.title_fill
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = self.border
        
        # CORREÇÃO: Aumentar altura da linha de cabeçalho
        ws.row_dimensions[start_row].height = header_height
        
        for row_idx, row_data in enumerate(dataframe_to_rows(df, index=False, header=False), start_row + 1):
            max_line_count = 1
            for col_idx, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.value = value
                cell.font = Font(name='Arial', size=data_font_size)
                cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
                cell.border = self.border
                
                # Contar linhas para ajustar altura
                if value:
                    line_count = str(value).count('\n') + 1
                    max_line_count = max(max_line_count, line_count)
                
                if (row_idx - start_row) % 2 == 0:
                    cell.fill = self.alt_row_fill
            
            # Ajustar altura da linha
            ws.row_dimensions[row_idx].height = max(18, max_line_count * 15)
        
        # Ajustar colunas com multiplicadores aumentados
        for col_idx, column_name in enumerate(df.columns, 1):
            header_length = len(str(column_name).replace('\n', ''))
            data_max = df[column_name].astype(str).str.len().max() if not df.empty else 10
            max_length = max(header_length * 1.4, data_max * 1.2)
            
            # Limites especiais para colunas conhecidas
            if 'Par' in str(column_name) and 'Crit' in str(column_name):
                col_width = min(max_length + 5, 100)
            elif 'ANSI' in str(column_name) or 'Cd.' in str(column_name):
                col_width = min(max_length + 3, 15)
            elif 'Tensao' in str(column_name) or 'Tensão' in str(column_name):
                col_width = min(max_length + 3, 18)
            else:
                col_width = min(max_length + 4, 70)
            
            ws.column_dimensions[self._get_column_letter(col_idx)].width = col_width
        
        # Rodapé
        footer_row = ws.max_row + 2
        footer = self.format_footer_text(report_title)
        ws.merge_cells(f'A{footer_row}:' + self._get_column_letter(len(df.columns)) + f'{footer_row}')
        footer_cell = ws[f'A{footer_row}']
        footer_cell.value = f"{footer['left']} | {footer['center']} | {footer['right']}"
        footer_cell.font = self.footer_font
        footer_cell.alignment = self.center_align
    
    @staticmethod
    def _get_column_letter(col_idx: int) -> str:
        """Converte índice de coluna para letra (1=A, 27=AA)"""
        result = ""
        while col_idx > 0:
            col_idx -= 1
            result = chr(col_idx % 26 + 65) + result
            col_idx //= 26
        return result
